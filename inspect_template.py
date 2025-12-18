import openpyxl
from pathlib import Path

template_path = Path.cwd() / "Formato Recibo" / "FORMATO DE INGRESOS DIARIOS.xlsx"
wb = openpyxl.load_workbook(template_path)
ws = wb.active

print(f"Max row: {ws.max_row}")
for row in range(1, ws.max_row + 1):
    values = []
    for col in range(1, 20):
        val = ws.cell(row=row, column=col).value
        if val:
            values.append(f"{ws.cell(row=row, column=col).coordinate}:{val}")
    if values:
        print(f"Row {row}: {values}")
