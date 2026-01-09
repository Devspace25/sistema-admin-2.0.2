from __future__ import annotations

from pathlib import Path
from datetime import datetime
import json
import os
import shutil


def _data_dir() -> Path:
    d = Path.cwd() / "data" / "receipts"
    d.mkdir(parents=True, exist_ok=True)
    return d


def print_receipt_80mm(*, order_code: str, product_name: str, total_usd: float, payment_method: str | None,
                       advisor: str | None, summary: str | None, details: dict | None) -> Path:
    """Genera un recibo de 80mm en texto plano. Devuelve la ruta del archivo creado.

    - order_code: p. ej., ORD-20250101-000123
    - details: parámetros del corpóreo (para auditoría)
    """
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    lines: list[str] = []
    lines.append("CENTRO DE IMPRESIÓN")
    lines.append("RIF: J-00000000-0")
    lines.append("TEL: 000-0000000")
    lines.append("-")
    lines.append(f"ORDEN: {order_code}")
    lines.append(f"FECHA: {now}")
    if advisor:
        lines.append(f"ASESOR: {advisor}")
    lines.append("-")
    lines.append(f"PRODUCTO: {product_name}")
    lines.append("-")
    pay = payment_method or ""
    lines.append(f"PAGO: {pay}")
    lines.append(f"TOTAL: $ {total_usd:.2f}")
    lines.append("-")
    lines.append("¡Gracias por su compra!")
    lines.append("")
    # Guardar
    out = _data_dir() / f"{order_code}.txt"
    out.write_text("\n".join(lines), encoding="utf-8")
    # Opcional: guardar JSON de detalles (auditoría)
    if details is not None:
        (out.with_suffix('.json')).write_text(json.dumps(details, ensure_ascii=False, indent=2), encoding='utf-8')
    return out


def _wrap_text(text: str, width: int = 32) -> list[str]:
    words = text.split()
    lines: list[str] = []
    cur: list[str] = []
    cur_len = 0
    for w in words:
        if cur_len + len(w) + (1 if cur else 0) > width:
            lines.append(" ".join(cur))
            cur = [w]
            cur_len = len(w)
        else:
            if cur:
                cur_len += 1 + len(w)
            else:
                cur_len = len(w)
            cur.append(w)
    if cur:
        lines.append(" ".join(cur))
    return lines


def _to_float(value: object, default: float = 0.0) -> float:
    """Convierte cualquier valor a float, devolviendo *default* si falla."""
    try:
        if value is None:
            return default
        if isinstance(value, str):
            cleaned = value.replace('.', '').replace(',', '.').strip()
            if cleaned:
                return float(cleaned)
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _format_bs(value: float) -> str:
    """Formatea un monto en bolívares con separador de miles y coma decimal."""
    amount = round(value or 0.0, 2)
    formatted = f"{amount:,.2f}"
    # Cambiar puntos por comas respetando separadores
    return formatted.replace(',', '¤').replace('.', ',').replace('¤', '.')


def print_order_80mm(*, order_id: int, sale_id: int, product_name: str, status: str, details_json: str) -> Path:
    """Genera un ticket de orden de producción (80mm) con detalles técnicos.

    - order_id: ID interno del pedido
    - sale_id: ID de la venta
    - product_name: nombre del producto
    - status: estado del pedido
    - details_json: JSON con los parámetros técnicos
    """
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    lines: list[str] = []
    lines.append("ORDEN DE PRODUCCIÓN")
    lines.append("-")
    lines.append(f"Pedido: #{order_id}")
    lines.append(f"Venta: #{sale_id}")
    lines.append(f"Fecha: {now}")
    lines.append(f"Producto: {product_name}")
    if status:
        lines.append(f"Estado: {status}")
    lines.append("-")
    # Detalles técnicos (estructura nueva esperada)
    try:
        details = json.loads(details_json or "{}")
    except Exception:
        details = {}

    # Try to load sale info (descripcion, forma_pago, diseno_usd, pagos)
    sale_info = {}
    try:
        from .db import make_session_factory
        from .models import Sale
        SF = make_session_factory()
        with SF() as ss:
            s_obj = ss.get(Sale, int(sale_id)) if sale_id else None
            if s_obj:
                sale_info = {
                    'descripcion': getattr(s_obj, 'descripcion', None),
                    'forma_pago': getattr(s_obj, 'forma_pago', None),
                    'diseno_usd': float(getattr(s_obj, 'diseno_usd', 0.0) or 0.0),
                    'ingresos_usd': float(getattr(s_obj, 'ingresos_usd', 0.0) or 0.0),
                    'abono_usd': float(getattr(s_obj, 'abono_usd', 0.0) or 0.0),
                    'venta_usd': float(getattr(s_obj, 'venta_usd', 0.0) or 0.0),
                    'restante': float(getattr(s_obj, 'restante', 0.0) or 0.0),
                }
    except Exception:
        sale_info = {}

    # If the details payload also contains descripcion_text or incluye_diseno, prefer them
    if 'descripcion_text' in details or 'items' in details:
        desc = details.get('descripcion_text') or ''
        if desc:
            lines.append("DESCRIPCION:")
            for l in _wrap_text(desc, 40):
                lines.append(l)
            lines.append("-")
        # override sale_info descripcion if present in details
        if desc:
            sale_info['descripcion'] = desc

        items = details.get('items') or []
        if items:
            lines.append("CANT  DESCRIPCION           SUBT$")
            for it in items:
                c = it.get('cantidad', 1)
                pu = it.get('precio_unitario', 0.0)
                sub = it.get('subtotal_usd', pu * c)
                lines.append(f"{c:>4}  {product_name[:18]:18}  ${sub:8.2f}")
            lines.append("-")

        totals = details.get('totals') or {}
        if totals:
            lines.append(f"TOTAL: $ {float(totals.get('total_usd', 0.0)):.2f}")
            lines.append(f"TOTAL Bs: {float(totals.get('total_bs', 0.0)):.2f}")
        # Add sale-level metadata: descripcion (if not yet added), forma de pago, incluye diseño, total pagado
        try:
            # Descripcion from sale_info if not already printed
            if not any(l.startswith('DESCRIPCION:') for l in lines):
                sdesc = sale_info.get('descripcion') or ''
                if sdesc:
                    lines.append('-')
                    lines.append('DESCRIPCION:')
                    for l in _wrap_text(str(sdesc), 40):
                        lines.append(l)
            # Forma de pago
            if sale_info.get('forma_pago'):
                lines.append(f"FORMA PAGO: {sale_info.get('forma_pago')}")
            # Incluye diseno
            incluye_diseno = False
            if ('incluye_diseno' in details and details.get('incluye_diseno')) or (sale_info.get('diseno_usd', 0.0) > 0.0):
                incluye_diseno = True
            lines.append(f"INCLUYE DISEÑO: {'SI' if incluye_diseno else 'NO'}")
            # Total pagado: prefer ingresos_usd, then abono_usd, then venta_usd-restante
            paid = None
            if sale_info.get('ingresos_usd') and sale_info.get('ingresos_usd') > 0:
                paid = sale_info.get('ingresos_usd')
            elif sale_info.get('abono_usd') and sale_info.get('abono_usd') > 0:
                paid = sale_info.get('abono_usd')
            elif sale_info.get('venta_usd') is not None:
                paid = float(sale_info.get('venta_usd') or 0.0) - float(sale_info.get('restante') or 0.0)
            if paid is not None:
                lines.append(f"TOTAL PAGADO: $ {float(paid):.2f}")
        except Exception:
            pass
    else:
        # Fallback: plano con cualquier clave conocida
        for key in [
            "alto_mm", "ancho_mm", "material_nombre", "espesor_mm",
            "tipo_corte", "base_tipo", "base_color",
            "supports", "luz", "regulador", "light_box_enabled", "light_box_pct",
            "subtotal", "total"
        ]:
            if key in details:
                val = details.get(key)
                lines.append(f"{key}: {val}")
    # Guardar archivo
    out = _data_dir() / f"ORDER-{int(order_id):06d}.txt"
    out.write_text("\n".join(lines), encoding="utf-8")
    # Guardar JSON completo al lado
    (out.with_suffix('.json')).write_text(json.dumps(details, ensure_ascii=False, indent=2), encoding='utf-8')
    return out


def print_order_pdf(*, order_id: int, sale_id: int, product_name: str, status: str, details_json: str, customer: dict | None = None, out_path: Path | None = None) -> Path:
    """Genera un ticket PDF de 80 mm con diseño profesional tipo recibo/factura."""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
        from reportlab.lib import colors
    except Exception:
        return print_order_80mm(order_id=order_id, sale_id=sale_id, product_name=product_name, status=status, details_json=details_json)

    out = Path(out_path) if out_path else (_data_dir() / f"ORDER-{int(order_id):06d}.pdf")

    try:
        details = json.loads(details_json or "{}")
    except Exception:
        details = {}

    items = details.get('items') if isinstance(details.get('items'), list) else []
    totals = details.get('totals') if isinstance(details.get('totals'), dict) else {}
    meta = details.get('meta') if isinstance(details.get('meta'), dict) else {}

    sale_info: dict[str, object] = {}
    customer_info: dict[str, str] = {}
    try:
        from .db import make_session_factory
        from .models import Sale, Customer

        SF = make_session_factory()
        with SF() as session:
            sale_obj = session.get(Sale, int(sale_id)) if sale_id else None
            if sale_obj:
                sale_info = {
                    'descripcion': getattr(sale_obj, 'descripcion', None),
                    'forma_pago': getattr(sale_obj, 'forma_pago', None),
                    'diseno_usd': _to_float(getattr(sale_obj, 'diseno_usd', 0.0)),
                    'ingresos_usd': _to_float(getattr(sale_obj, 'ingresos_usd', 0.0)),
                    'abono_usd': _to_float(getattr(sale_obj, 'abono_usd', 0.0)),
                    'venta_usd': _to_float(getattr(sale_obj, 'venta_usd', 0.0)),
                    'restante': _to_float(getattr(sale_obj, 'restante', 0.0)),
                    'total_bs': _to_float(getattr(sale_obj, 'total_bs', getattr(sale_obj, 'monto_bs', 0.0))),
                    'monto_bs': _to_float(getattr(sale_obj, 'monto_bs', 0.0)),
                    'tasa_bcv': _to_float(getattr(sale_obj, 'tasa_bcv', None), default=0.0),
                    'iva': _to_float(getattr(sale_obj, 'iva', 0.0)),
                    'fecha': getattr(sale_obj, 'fecha', None),
                    'fecha_pago': getattr(sale_obj, 'fecha_pago', None),
                    'asesor': getattr(sale_obj, 'asesor', None),
                    'numero_orden': getattr(sale_obj, 'numero_orden', None),
                    'cliente': getattr(sale_obj, 'cliente', None),
                    'cliente_id': getattr(sale_obj, 'cliente_id', None),
                    'notes': getattr(sale_obj, 'notes', None),
                }
                raw_cid = getattr(sale_obj, 'cliente_id', None)
                if raw_cid:
                    try:
                        cust_obj = session.get(Customer, int(raw_cid))
                    except Exception:
                        cust_obj = None
                    if cust_obj:
                        customer_info = {
                            'name': cust_obj.name or ((cust_obj.first_name or '') + ' ' + (cust_obj.last_name or '')).strip(),
                            'short_address': cust_obj.short_address or '',
                            'document': cust_obj.document or '',
                            'phone': cust_obj.phone or '',
                        }
    except Exception:
        sale_info = sale_info or {}

    # Priorizar customer recibido, luego meta, luego sale
    resolved_customer = {
        'name': '',
        'short_address': '',
        'document': '',
        'phone': '',
    }
    if isinstance(customer, dict):
        resolved_customer.update({k: str(customer.get(k) or '') for k in resolved_customer.keys()})

    if isinstance(meta.get('cliente'), dict):
        resolved_customer.update({k: str(meta['cliente'].get(k) or '') for k in resolved_customer.keys() if isinstance(meta['cliente'], dict)})
    elif isinstance(meta.get('cliente'), str) and not resolved_customer['name']:
        resolved_customer['name'] = str(meta.get('cliente'))

    if not resolved_customer['name'] and sale_info.get('cliente'):
        resolved_customer['name'] = str(sale_info.get('cliente') or '')

    # Merge customer info from DB last to avoid overwriting explicit data with empties
    for key, value in customer_info.items():
        if value and not resolved_customer.get(key):
            resolved_customer[key] = value

    # Complement metadata from meta keys if available
    for alias_key in ('documento', 'rif', 'rif_ci', 'identificacion'):
        if not resolved_customer['document'] and isinstance(meta.get(alias_key), str):
            resolved_customer['document'] = meta[alias_key]

    if not resolved_customer['short_address'] and isinstance(meta.get('direccion'), str):
        resolved_customer['short_address'] = meta.get('direccion') or ''

    if not resolved_customer['phone'] and isinstance(meta.get('telefono'), str):
        resolved_customer['phone'] = meta.get('telefono') or ''

    rate_bcv = _to_float(sale_info.get('tasa_bcv'), 0.0) or _to_float(meta.get('tasa_bcv'), 0.0)
    total_bs = _to_float(totals.get('total_bs'), 0.0)
    total_usd = _to_float(totals.get('total_usd'), 0.0)
    if total_bs == 0.0 and rate_bcv and total_usd:
        total_bs = total_usd * rate_bcv
    if total_bs == 0.0:
        total_bs = _to_float(sale_info.get('total_bs') or sale_info.get('monto_bs'), 0.0)

    diseno_usd = _to_float(sale_info.get('diseno_usd'), 0.0)
    diseno_bs = diseno_usd * rate_bcv if rate_bcv else 0.0
    if diseno_bs == 0.0:
        diseno_bs = _to_float(meta.get('diseno_bs'), 0.0)

    instalacion_bs = _to_float(meta.get('instalacion_bs') or meta.get('instalacion'), 0.0)
    iva_amount = _to_float(sale_info.get('iva'), 0.0)

    description_text = str(details.get('descripcion_text') or sale_info.get('descripcion') or product_name or '').strip()
    
    # Detectar si la descripción es redundante (concatenación de items)
    if items and description_text:
        try:
            item_descs = []
            for item in items:
                d = str(item.get('descripcion') or item.get('label') or item.get('product_name') or '').strip()
                if d:
                    item_descs.append(d)
            auto_desc = " + ".join(item_descs)
            # Si es idéntica, la suprimimos para no duplicar info en el ticket
            if description_text.strip() == auto_desc.strip():
                description_text = ""
        except Exception:
            pass

    description_lines: list[str] = []
    
    if description_text:
        for raw_line in description_text.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            description_lines.extend(_wrap_text(line, width=38))

    # Agregar notas adicionales si existen
    notes_text = str(sale_info.get('notes') or '').strip()
    if notes_text:
        for raw_line in notes_text.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            description_lines.extend(_wrap_text(line, width=38))

    # Solo usar fallback si no hay nada que mostrar y no hay items
    if not description_lines and not items:
        description_lines.append(product_name or '—')

    # Preparar datos de items (descripcion, cantidad y subtotal en Bs)
    item_lines: list[tuple[str, float, float]] = []
    if items:
        for idx, item in enumerate(items, start=1):
            qty = _to_float(item.get('cantidad'), 0.0) or 0.0
            subtotal_bs = _to_float(item.get('subtotal_bs'), 0.0)
            if subtotal_bs == 0.0 and rate_bcv:
                subtotal_bs = _to_float(item.get('subtotal_usd'), 0.0) * rate_bcv
            if subtotal_bs == 0.0 and idx == 1 and total_bs:
                subtotal_bs = total_bs
            
            # Priorizar descripción específica del item
            raw_desc = str(item.get('descripcion') or item.get('product_name') or item.get('label') or product_name or f'Item {idx}')
            
            # Si la descripción contiene " + " y es muy larga, probablemente es la concatenación global.
            # Intentar limpiar si es posible, o usar solo la primera parte si coincide con el nombre del producto.
            if " + " in raw_desc and len(raw_desc) > 50:
                # Si tenemos un nombre de producto específico en el item, úsalo
                if item.get('product_name'):
                    raw_desc = item.get('product_name')
                else:
                    # Si no, intentar tomar el primer segmento
                    parts = raw_desc.split(" + ")
                    if parts:
                        raw_desc = parts[0]

            item_lines.append((raw_desc, qty, subtotal_bs))
    else:
        item_lines.append((product_name or 'Servicio', 1.0, total_bs or summary_amount))

    # Detectar desglose de pagos desde el payload/meta (mover después de item_lines para evitar referencias previas)
    payment_lines: list[tuple[str, str, float]] = []
    raw_payments = None
    for key in ('payment_breakdown', 'payments', 'metodos_pago'):
        if isinstance(meta.get(key), (list, dict)):
            raw_payments = meta.get(key)
            break
    if raw_payments:
        if isinstance(raw_payments, list):
            for entry in raw_payments:
                if not isinstance(entry, dict):
                    continue
                method = str(entry.get('method') or entry.get('metodo') or '')
                concept = str(entry.get('label') or entry.get('concepto') or method)
                amount_bs = _to_float(entry.get('amount_bs') or entry.get('monto_bs') or entry.get('monto'))
                payment_lines.append((method or concept, concept, amount_bs))
        elif isinstance(raw_payments, dict):
            for method, entry in raw_payments.items():
                if isinstance(entry, dict):
                    concept = str(entry.get('label') or entry.get('concepto') or method)
                    amount_bs = _to_float(entry.get('amount_bs') or entry.get('monto_bs') or entry.get('monto'))
                else:
                    concept = str(method)
                    amount_bs = _to_float(entry)
                payment_lines.append((str(method), concept, amount_bs))

    if not payment_lines:
        primary_method = str(sale_info.get('forma_pago') or meta.get('forma_pago') or 'Efectivo$')
        payment_lines = [
            (primary_method, 'Total Venta', total_bs),
            ('Zelle$', 'Diseño', diseno_bs),
            ('Pago Móvil', 'Instalación', instalacion_bs),
            ('Punto Venta', 'I.V.A 16%', iva_amount),
        ]

    # Calcular totales derivados tras conocer item_lines
    summary_amount = sum(amount for _, _, amount in payment_lines if amount)
    items_total_bs = sum(amount for _, _, amount in item_lines)
    total_bs_display = total_bs or summary_amount

    # Resolver códigos visibles
    order_code = str(meta.get('order_number') or meta.get('numero_orden') or f"ORD-{int(order_id):06d}")
    order_digits = ''.join(ch for ch in order_code if ch.isdigit()) or f"{int(order_id):06d}"
    try:
        from .repository import get_sale_display_number
        sale_display = get_sale_display_number(int(sale_id), payment_date=sale_info.get('fecha_pago')) if sale_id else ''
    except Exception:
        sale_display = f"V-{int(sale_id):06d}" if sale_id else ''

    entrega_label = ''
    for key in ('fecha_entrega', 'delivery_date', 'entrega'):
        if meta.get(key):
            entrega_label = str(meta.get(key))
            break
    if not entrega_label:
        entrega_label = 'N/A'

    asesor_label = str(sale_info.get('asesor') or meta.get('asesor') or '').strip()
    if not asesor_label and resolved_customer['name']:
        asesor_label = resolved_customer['name']

    emission_time = datetime.now().strftime('%H:%M:%S')

    # --- Diseño profesional tipo recibo/factura (REDISEÑO SIMPLE) ---
    # Asegurar dimensiones de 80mm
    width = 80 * mm
    # Configuración específica solicitada: 80x297mm
    page_height = 297 * mm 
    
    c = canvas.Canvas(str(out), pagesize=(width, page_height))
    c.setPageSize((width, page_height))
    
    # Márgenes y espaciado
    margin_x = 2 * mm 
    margin_y = 1 * mm 
    line_h = 4.0 * mm 
    y = page_height - margin_y

    def draw_centered(text: str, y_pos: float, font: str = 'Helvetica', size: int = 9, bold: bool = False) -> float:
        font_name = f'{font}-Bold' if bold else font
        c.setFont(font_name, size)
        c.drawCentredString(width / 2, y_pos, text)
        return y_pos - line_h

    def draw_left(text: str, y_pos: float, x_pos: float = None, font: str = 'Helvetica', size: int = 8, bold: bool = False) -> float:
        font_name = f'{font}-Bold' if bold else font
        c.setFont(font_name, size)
        c.drawString(x_pos if x_pos is not None else margin_x, y_pos, text)
        return y_pos - line_h

    def draw_right(text: str, y_pos: float, font: str = 'Helvetica', size: int = 8, bold: bool = False) -> float:
        font_name = f'{font}-Bold' if bold else font
        c.setFont(font_name, size)
        c.drawRightString(width - margin_x, y_pos, text)
        return y_pos - line_h

    def draw_line_separator(y_pos: float, width_line: float = 1.5) -> float:
        c.setLineWidth(width_line)
        c.line(margin_x, y_pos, width - margin_x, y_pos)
        return y_pos - line_h * 0.5

    # ============================================================
    # ENCABEZADO (LOGO + DATOS)
    # ============================================================
    # Logo placeholder
    logo_path = Path(__file__).parent.parent.parent / 'assets' / 'img' / 'logo.png'
    if logo_path.exists():
        # Centered logo
        img_w = 30 * mm
        img_h = 10 * mm
        c.drawImage(str(logo_path), (width - img_w)/2, y - img_h, width=img_w, height=img_h, preserveAspectRatio=True, mask='auto')
        y -= img_h + 2

    y = draw_centered('MR.7 PUBLICIDAD, C.A. J-506410990', y, size=10, bold=True)
    y = draw_centered('Av. Universidad, Centro Parque Carabobo, Torre A, Piso 4, Ofic 413', y, size=7)
    y -= line_h * 0.5

    # ============================================================
    # DATOS CLIENTE / ORDEN (2 COLUMNAS)
    # ============================================================
    # Columna izquierda (Labels + Values)
    left_x = margin_x
    # Columna derecha (Labels + Values aligned right)
    right_x = width - margin_x

    # Fila 1: Cliente | Orden
    c.setFont('Helvetica-Bold', 8)
    c.drawString(left_x, y, 'Cliente:')
    c.setFont('Helvetica', 8)
    # Truncar nombre cliente si es muy largo
    c.drawString(left_x + 12*mm, y, (resolved_customer.get('name') or '---Seleccione---')[:20])
    
    c.setFont('Helvetica-Bold', 8)
    c.drawRightString(right_x, y, 'ORDEN Nº')
    y -= line_h

    # Fila 2: Dirección | Numero Orden
    c.setFont('Helvetica-Bold', 8)
    c.drawString(left_x, y, 'Dirección:')
    c.setFont('Helvetica', 7) # Address smaller
    addr = resolved_customer.get('short_address') or ''
    c.drawString(left_x + 14*mm, y, addr[:25])
    
    c.setFont('Helvetica-Bold', 10)
    c.drawRightString(right_x, y, order_digits.zfill(6))
    y -= line_h

    # Fila 3: RIF | Fecha
    c.setFont('Helvetica-Bold', 8)
    c.drawString(left_x, y, 'R.I.F/C.I:')
    c.setFont('Helvetica', 8)
    c.drawString(left_x + 14*mm, y, resolved_customer.get('document') or '')
    
    c.setFont('Helvetica', 8)
    c.drawRightString(right_x, y, datetime.now().strftime("%d/%m/%Y"))
    y -= line_h

    # Fila 4: Telefono | Hora
    c.setFont('Helvetica-Bold', 8)
    c.drawString(left_x, y, 'Teléfono:')
    c.setFont('Helvetica', 8)
    c.drawString(left_x + 14*mm, y, resolved_customer.get('phone') or '')
    
    c.setFont('Helvetica', 8)
    c.drawRightString(right_x, y, emission_time)
    y -= line_h * 0.5

    # ============================================================
    # TABLA ITEMS
    # ============================================================
    # Header Line
    y = draw_line_separator(y, width_line=1.5)
    
    # Headers
    c.setFont('Helvetica-Bold', 8)
    col_cant = margin_x + 2
    col_desc = margin_x + 12
    col_total = width - margin_x - 2
    
    c.drawString(col_cant, y, 'Cant')
    c.drawCentredString(width/2, y, 'Descripción')
    c.drawRightString(col_total, y, 'Total Bs')
    y -= line_h * 0.8
    
    # Separator
    y = draw_line_separator(y, width_line=1.5)
    y -= line_h * 0.2

    # Items Loop
    c.setFont('Helvetica', 8)
    for desc, qty, amount in item_lines:
        qty_str = f"{int(qty)}" if abs(qty - int(qty)) < 1e-6 else f"{qty:.1f}"
        
        lines = _wrap_text(desc, width=35) 
        
        c.drawString(col_cant, y, qty_str.rjust(3))
        c.drawString(col_desc, y, lines[0] if lines else "")
        c.drawRightString(col_total, y, _format_bs(amount))
        
        y -= line_h
        
        if len(lines) > 1:
            c.setFont('Helvetica', 7)
            for extra_line in lines[1:]:
                c.drawString(col_desc, y, extra_line)
                y -= line_h * 0.8
            c.setFont('Helvetica', 8)
            
        y -= line_h * 0.3 
    
    y -= line_h * 0.2
    y = draw_line_separator(y, width_line=1.5)
    y -= line_h * 0.2

    # ============================================================
    # TOTALES
    # ============================================================
    # M.P. (Metodo Pago)
    c.setFont('Helvetica-Bold', 8)
    c.drawString(left_x, y, 'M.P.:')
    # Draw line under M.P.
    c.setLineWidth(1)
    c.line(left_x, y - 2, left_x + 15*mm, y - 2)
    
    # Totals Right Aligned
    val_x = width - margin_x
    lbl_x = val_x - 25*mm

    c.setFont('Helvetica', 8)
    c.drawRightString(lbl_x, y, 'Subtotal')
    c.drawRightString(val_x, y, _format_bs(items_total_bs))
    y -= line_h

    if diseno_bs > 0:
        c.drawRightString(lbl_x, y, 'Diseño')
        c.drawRightString(val_x, y, _format_bs(diseno_bs))
        y -= line_h

    if instalacion_bs > 0:
        c.drawRightString(lbl_x, y, 'Insts.')
        c.drawRightString(val_x, y, _format_bs(instalacion_bs))
        y -= line_h

    if iva_amount > 0:
        c.drawRightString(lbl_x, y, 'I.V.A 16%')
        c.drawRightString(val_x, y, _format_bs(iva_amount))
        y -= line_h

    # Separator for Total
    c.setLineWidth(1)
    c.line(lbl_x - 5, y + 2, val_x, y + 2)
    
    # Asesor (Left) & Total (Right)
    c.setFont('Helvetica-Bold', 9)
    c.drawString(left_x, y, f'ASESOR: {asesor_label[:15]}')
    # Line under Asesor
    c.line(left_x, y - 2, left_x + 30*mm, y - 2)

    c.drawRightString(lbl_x, y, 'TOTAL Bs')
    c.drawRightString(val_x, y, _format_bs(total_bs_display))
    y -= line_h * 1.5

    # ============================================================
    # FOOTER
    # ============================================================
    y = draw_centered('INFORMACION IMPORTANTE', y, size=8, bold=True)
    
    c.setFont('Helvetica', 6)
    footer_lines = [
        'Estimado Cliente: Por favor verifique su pedido antes de retirarse.',
        'Bajo ningún concepto se realizan devoluciones por compras erróneas.',
    ]
    for line in footer_lines:
        y = draw_centered(line, y, size=6)
    
    y = draw_line_separator(y, width_line=1.5)
    y = draw_centered('Este Documento no da derecho a crédito fiscal. Nuestros precios no incluyen I.V.A', y, size=6, bold=True)
    y -= line_h * 0.5
    
    y = draw_centered('EXIJA SU FACTURA FISCAL', y, size=7)
    y -= line_h * 0.5

    # ABONO / POR COBRAR
    abono_val = sale_info.get('abono_usd', 0.0)
    restante_val = sale_info.get('restante', 0.0)
    
    footer_text = f"ABONO $ {abono_val:.2f} POR COBRAR $ {restante_val:.2f}"
    y = draw_centered(footer_text, y, size=11, bold=True)

    c.showPage()
    c.save()
    (out.with_suffix('.json')).write_text(json.dumps(details, ensure_ascii=False, indent=2), encoding='utf-8')
    return out


def print_ticket_excel_pdf(order_info: dict, output_path: Path = None) -> Path:
    """
    Rellena el template de Excel y lo exporta a PDF.
    order_info debe contener:
    - order_number
    - date
    - customer_name, customer_address, customer_rif, customer_phone
    - items: list of dict(qty, desc, total)
    - subtotal, tax, total
    - advisor
    - payment_method
    """
    import openpyxl
    try:
        import win32com.client
        import pythoncom
    except ImportError:
        win32com = None

    template_path = Path.cwd() / "Formato Recibo" / "formato_recibo.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"No se encontró el formato en: {template_path}")

    # Crear archivo temporal para el Excel lleno
    temp_dir = Path.cwd() / "data" / "temp"
    temp_dir.mkdir(parents=True, exist_ok=True)
    temp_excel = temp_dir / f"temp_ticket_{order_info.get('order_number', 'new')}.xlsx"
    
    shutil.copy2(template_path, temp_excel)
    
    wb = openpyxl.load_workbook(temp_excel)
    ws = wb.active # Asumimos que es la hoja activa o "TICKET"
    
    # Mapeo de datos
    # Encabezados
    # C3: Cliente
    if order_info.get('customer_name'):
        ws['C3'] = order_info['customer_name']
    # C4: Dirección
    if order_info.get('customer_address'):
        ws['C4'] = order_info['customer_address']
    # C5: RIF
    if order_info.get('customer_rif'):
        ws['C5'] = order_info['customer_rif']
    # C6: Teléfono
    if order_info.get('customer_phone'):
        ws['C6'] = order_info['customer_phone']
        
    # E4: Orden No
    if order_info.get('order_number'):
        ws['E4'] = order_info['order_number']
        
    # Items (Desde fila 9)
    items = order_info.get('items', [])
    start_row = 9
    max_rows = 6 # Hasta fila 14 aprox
    
    for i, item in enumerate(items[:max_rows]):
        row = start_row + i
        ws[f'A{row}'] = item.get('qty', 1)
        ws[f'B{row}'] = item.get('desc', '')
        ws[f'E{row}'] = item.get('total', 0.0)
        
    # Totales
    # B14: Metodo Pago (M.P.)
    if order_info.get('payment_method'):
        ws['B14'] = order_info['payment_method']
        
    # E15: Diseño (Bs)
    if order_info.get('design_bs'):
        ws['E15'] = order_info['design_bs']
        
    # E16: Instalación (Bs)
    if order_info.get('installation_bs'):
        ws['E16'] = order_info['installation_bs']
        
    # E17: IVA (Bs)
    if order_info.get('iva_bs'):
        ws['E17'] = order_info['iva_bs']
        
    # C18: Asesor
    if order_info.get('advisor'):
        ws['C18'] = order_info['advisor']
        
    # A24: Estado de Pago (PAGO TOTAL o ABONO/RESTANTE)
    if order_info.get('payment_status_text'):
        ws['A24'] = order_info['payment_status_text']
        
    wb.save(temp_excel)
    wb.close()
    
    # Convertir a PDF usando Excel COM
    if output_path is None:
        output_path = Path.cwd() / "data" / "receipts" / f"ticket_{order_info.get('order_number', 'temp')}.pdf"
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
    _excel_to_pdf(temp_excel, output_path)
    
    # Limpiar temporal
    try:
        os.remove(temp_excel)
    except:
        pass
        
    return output_path

def _excel_to_pdf(excel_path: Path, pdf_path: Path):
    import win32com.client
    import pythoncom
    import win32print
    
    pythoncom.CoInitialize()
    
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    
    try:
        # Intentar cambiar a una impresora que soporte Carta (Microsoft Print to PDF)
        # Esto evita que Excel use la configuración de la impresora térmica (80mm)
        # target_printer = "Microsoft Print to PDF"
        # printer_set = False
        
        # COMENTADO: Si queremos respetar el formato del Excel (que ya está en 80mm),
        # no deberíamos forzar una impresora que usa Carta por defecto.
        # Dejamos que use la impresora predeterminada (que podría ser la térmica)
        # o simplemente confiamos en que ExportAsFixedFormat use la config de página del Excel.
        pass
        
        wb = excel.Workbooks.Open(str(excel_path.resolve()))
        
        # Forzar configuración de página vía COM
        ws = wb.ActiveSheet
        try:
            # No forzar tamaño carta (xlPaperLetter = 1) si queremos 80mm
            # Si el Excel ya tiene el tamaño correcto, no deberíamos tocar PaperSize
            # O si queremos forzarlo, deberíamos buscar el ID de papel custom o similar.
            # Pero el usuario dice que el Excel ya es el formato.
            
            # Comentamos PaperSize para respetar el del archivo
            # ws.PageSetup.PaperSize = 1 
            
            # Orientación: xlPortrait = 1, xlLandscape = 2
            # Para ticket suele ser Portrait
            # ws.PageSetup.Orientation = 1
            
            # Zoom = False para permitir FitToPages
            ws.PageSetup.Zoom = False
            ws.PageSetup.FitToPagesWide = 1
            ws.PageSetup.FitToPagesTall = False # Permitir que crezca verticalmente
            
            # Márgenes (en puntos, 1 cm ~= 28.35 pt)
            # Respetar márgenes del Excel si es posible, o poner mínimos
            # ws.PageSetup.LeftMargin = 0
            # ws.PageSetup.RightMargin = 0
            # ws.PageSetup.TopMargin = 0
            # ws.PageSetup.BottomMargin = 0
            # ws.PageSetup.HeaderMargin = 0
            # ws.PageSetup.FooterMargin = 0
            # ws.PageSetup.CenterHorizontally = True
            pass
            
        except Exception:
            pass

        # ExportAsFixedFormat 0 = xlTypePDF
        wb.ExportAsFixedFormat(0, str(pdf_path.resolve()))
        wb.Close(False)
    except Exception as e:
        raise e
    finally:
        excel.Quit()


def print_daily_report_excel_pdf(sales_data: list[dict], report_date: datetime) -> Path:
    """
    Genera un reporte diario en PDF usando la plantilla Excel.
    Nueva implementación optimizada para paginación correcta.
    """
    import openpyxl
    import shutil
    from copy import copy
    
    template_path = Path.cwd() / "Formato Recibo" / "FORMATO DE INGRESOS DIARIOS.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"No se encontró la plantilla en: {template_path}")
        
    # Crear archivo temporal
    temp_excel = _data_dir() / f"temp_report_{int(datetime.now().timestamp())}.xlsx"
    shutil.copy2(template_path, temp_excel)
    
    wb = openpyxl.load_workbook(temp_excel)
    ws = wb.active
    
    # 1. Configuración de Página (CRÍTICO para que salga en 1 hoja)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = 1  # Letter / Carta
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.scale = None  # Importante: desactivar escala manual
    
    # Márgenes mínimos
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.1
    ws.page_margins.bottom = 0.1
    ws.page_margins.header = 0.0
    ws.page_margins.footer = 0.0
    
    ws.print_options.horizontalCentered = True
    
    # Limpiar saltos de página manuales de la plantilla
    try:
        # openpyxl RowBreak object doesn't have clear(), access internal list if needed
        if hasattr(ws, 'row_breaks') and hasattr(ws.row_breaks, 'brk'):
            ws.row_breaks.brk = []
        if hasattr(ws, 'col_breaks') and hasattr(ws.col_breaks, 'brk'):
            ws.col_breaks.brk = []
    except Exception:
        pass

    # 2. Llenar Fecha del Reporte (Sobrescribir fórmulas TODAY())
    # Según inspección: O4=Día, P4=Mes, Q4=Año
    ws['O4'] = report_date.day
    ws['P4'] = report_date.month
    ws['Q4'] = report_date.year

    # 3. Gestión de Filas Dinámicas
    # La plantilla tiene espacio para 2 ventas (filas 8 y 9).
    # Si hay más, insertamos filas entre la 8 y la 9.
    start_row = 8
    num_sales = len(sales_data)
    
    if num_sales > 2:
        rows_to_add = num_sales - 2
        # Insertar antes de la fila 9 (la segunda fila de datos original)
        ws.insert_rows(9, amount=rows_to_add)
        
        # Copiar estilos de la fila 8 a las nuevas filas
        for r_idx in range(9, 9 + rows_to_add):
            for c_idx in range(1, 20): # Columnas A-S
                source = ws.cell(row=8, column=c_idx)
                target = ws.cell(row=r_idx, column=c_idx)
                if source.has_style:
                    target.font = copy(source.font)
                    target.border = copy(source.border)
                    target.fill = copy(source.fill)
                    target.number_format = copy(source.number_format)
                    target.alignment = copy(source.alignment)
                    target.protection = copy(source.protection)

    # 4. Llenar Datos
    for i, sale in enumerate(sales_data):
        row = start_row + i
        
        # Mapeo de columnas
        ws[f'A{row}'] = sale.get('numero_orden', '')
        ws[f'B{row}'] = sale.get('articulo', '')
        ws[f'C{row}'] = sale.get('asesor', '')
        ws[f'D{row}'] = sale.get('venta_usd', 0.0)
        ws[f'E{row}'] = sale.get('forma_pago', '')
        ws[f'F{row}'] = sale.get('serial_billete', '')
        ws[f'G{row}'] = sale.get('banco', '')
        ws[f'H{row}'] = sale.get('referencia', '')
        
        # Fecha pago
        fp = sale.get('fecha_pago')
        ws[f'I{row}'] = str(fp)[:10] if fp else ''
        
        ws[f'J{row}'] = sale.get('monto_bs', 0.0)
        ws[f'K{row}'] = sale.get('monto_usd_calculado', 0.0)
        ws[f'L{row}'] = sale.get('abono_usd', 0.0)
        ws[f'M{row}'] = "" # Restante $ (Placeholder para cobros futuros)
        ws[f'N{row}'] = sale.get('iva', 0.0)
        ws[f'O{row}'] = sale.get('restante', 0.0) # Por cobrar
        ws[f'P{row}'] = sale.get('diseno_usd', 0.0)
        ws[f'Q{row}'] = 0.0 # Instalación
        ws[f'S{row}'] = sale.get('ingresos_usd', 0.0)

    # 5. Definir Área de Impresión
    # Desde A1 hasta la última fila con contenido (incluyendo firmas)
    ws.print_area = f'A1:S{ws.max_row}'
    
    wb.save(temp_excel)
    wb.close()
    
    # 6. Convertir a PDF
    output_path = _data_dir() / f"Reporte_Diario_{report_date.strftime('%Y-%m-%d')}.pdf"
    _excel_to_pdf(temp_excel, output_path)
    
    try:
        os.remove(temp_excel)
    except:
        pass
        
    return output_path

def _print_daily_report_excel_pdf_old(sales_data: list[dict], report_date: datetime) -> Path:
    """Genera un reporte diario en PDF usando la plantilla Excel."""
    import openpyxl
    import shutil
    
    template_path = Path.cwd() / "Formato Recibo" / "FORMATO DE INGRESOS DIARIOS.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"No se encontró la plantilla en: {template_path}")
        
    # Crear temporal
    temp_excel = _data_dir() / f"temp_report_{int(datetime.now().timestamp())}.xlsx"
    shutil.copy2(template_path, temp_excel)
    
    wb = openpyxl.load_workbook(temp_excel)
    ws = wb.active
    
    # Configuración de página según captura (Carta, Horizontal)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = 1  # Carta / Letter (1 = Letter, 9 = A4)
    
    # FORZAR AJUSTE A UNA PÁGINA
    # Aunque la captura muestra 100%, si el contenido es más ancho/alto que la página, se dividirá.
    # Para garantizar 1 sola página, usamos fitToPage.
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.scale = None # Desactivar escala manual para que fitToPage funcione
    
    # Márgenes según captura
    ws.page_margins.top = 0.1
    ws.page_margins.bottom = 0.1
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.0
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1
    
    # Centrar horizontalmente
    ws.print_options.horizontalCentered = True
    
    # Eliminar saltos de página existentes que puedan causar división
    try:
        # En versiones recientes de openpyxl, se usa row_breaks y col_breaks
        if hasattr(ws, 'row_breaks'):
            ws.row_breaks.clear()
        if hasattr(ws, 'col_breaks'):
            ws.col_breaks.clear()
    except Exception:
        pass
    
    # Llenar fecha del reporte (asumiendo que hay un lugar para ello, si no, lo ignoramos o lo ponemos en A1)
    # El usuario no especificó celda para fecha del reporte, pero es buena práctica.
    # Revisando instrucciones: "el orden de los datos son a partir de la fila 8"
    # Asumiremos que la fecha puede ir en el título o similar, pero sin instrucción específica, solo llenamos la tabla.
    
    start_row = 8
    from copy import copy
    
    # Función helper para copiar estilos
    def copy_style(source_cell, target_cell):
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

    # Lógica de inserción de filas:
    # La plantilla tiene 2 filas de datos predefinidas (8 y 9).
    # Si hay más de 2 ventas, insertamos filas adicionales ENTRE la 8 y la 9 para expandir el rango de suma.
    num_sales = len(sales_data)
    if num_sales > 2:
        rows_to_insert = num_sales - 2
        # Insertar en la fila 9. Esto empuja la fila 9 original hacia abajo.
        # El rango de suma D8:D9 se expandirá automáticamente.
        ws.insert_rows(9, amount=rows_to_insert)
        
        # Copiar estilos de la fila 8 a las nuevas filas insertadas
        for r in range(9, 9 + rows_to_insert):
            for col in range(1, 20): # A hasta S
                source_cell = ws.cell(row=8, column=col)
                target_cell = ws.cell(row=r, column=col)
                copy_style(source_cell, target_cell)

    # Llenar datos
    for i, sale in enumerate(sales_data):
        row = start_row + i
        
        # A: Numero de orden
        ws[f'A{row}'] = sale.get('numero_orden', '')
        
        # B: Producto
        ws[f'B{row}'] = sale.get('articulo', '')
        
        # C: Asesor
        ws[f'C{row}'] = sale.get('asesor', '')
        
        # D: Venta $
        ws[f'D{row}'] = sale.get('venta_usd', 0.0)
        
        # E: Forma de pago
        ws[f'E{row}'] = sale.get('forma_pago', '')
        
        # F: Serial del billete
        ws[f'F{row}'] = sale.get('serial_billete', '')
        
        # G: Banco
        ws[f'G{row}'] = sale.get('banco', '')
        
        # H: Referencia bancaria
        ws[f'H{row}'] = sale.get('referencia', '')
        
        # I: Fecha de pago
        fecha_pago = sale.get('fecha_pago')
        if fecha_pago:
            # Si es string ISO, cortar fecha
            if isinstance(fecha_pago, str):
                ws[f'I{row}'] = fecha_pago[:10]
            else:
                ws[f'I{row}'] = str(fecha_pago)
        
        # J: Monto Bs
        ws[f'J{row}'] = sale.get('monto_bs', 0.0)
        
        # K: Monto $ (Calculado)
        ws[f'K{row}'] = sale.get('monto_usd_calculado', 0.0)
        
        # L: Abono $
        ws[f'L{row}'] = sale.get('abono_usd', 0.0)
        
        # M: Restante $
        ws[f'M{row}'] = sale.get('restante', 0.0)
        
        # N: IVA
        ws[f'N{row}'] = sale.get('iva', 0.0)
        
        # O: Por Cobrar $ (Asumimos Restante)
        ws[f'O{row}'] = sale.get('restante', 0.0)
        
        # P: Diseño
        ws[f'P{row}'] = sale.get('diseno_usd', 0.0)
        
        # Q: Inst (Instalación - No tenemos campo, ponemos 0)
        ws[f'Q{row}'] = 0.0
        
        # S: Ingresos $ (Saltamos R)
        ws[f'S{row}'] = sale.get('ingresos_usd', 0.0)
        
    # Definir área de impresión: Todo el contenido hasta la última fila usada
    # Esto asegura que se incluyan los totales y firmas, sin cortar nada.
    ws.print_area = f'A1:S{ws.max_row}'
        
    wb.save(temp_excel)
    wb.close()
    
    # Generar PDF
    output_path = _data_dir() / f"Reporte_Diario_{report_date.strftime('%Y-%m-%d')}.pdf"
    _excel_to_pdf(temp_excel, output_path)
    
    # Limpiar
    try:
        os.remove(temp_excel)
    except:
        pass
        
    return output_path
